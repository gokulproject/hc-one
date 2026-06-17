"""
ExcelCreation/excel_manager.py
==============================
Generate Excel report for Oracle DB Connection Check.

Sheet 1 — DB Connection Check (unchanged columns, two new columns added):
  # | Customer | Environment | Server Name | Server ID | Server Host |
  DB Name | DB Type | DSN | DB Connection Status |
  Duration (ms) | Error | Error Type | Checked At (IST) |
  Expired (Action Required) | Reason (if yes)

Sheet 2 — Customer - Environment Report (FMW only):
  Customer | Environment | Server Host / DB Name / DB Type |
  USERNAME | ACCOUNT_STATUS | EXPIRY_DATE

Expiry validation rules (Customer/Environment level, FMW only):
  - Any OAM account with a non-null/non-empty EXPIRY_DATE  →  Expired=Yes, Action Required=Yes
  - All accounts have null/none/empty EXPIRY_DATE           →  Expired=No,  Action Required=No
  - Non-FMW db_type: columns left blank
"""
import logging
import os
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log      = logging.getLogger("checker.excel")
IST      = timedelta(hours=5, minutes=30)
LAST_COL = 16   # 14 original + 2 new expiry columns

# ── Palette ───────────────────────────────────────────────────────────────────
HDR_DARK   = "1565C0"
HDR_MED    = "0D47A1"
HDR_GREEN  = "1B5E20"   # header for Customer-Env Report sheet
SUMM_BG    = "E3F2FD"
GREEN_BG   = "E8F5E9";  GREEN_BG2  = "F1F8E9"
RED_BG     = "FFEBEE";  RED_BG2    = "FFF5F5"
AMBER_BG   = "FFF8E1";  AMBER_BG2  = "FFFDE7"
C_GREEN    = "2E7D32"
C_RED      = "C62828"
C_AMBER    = "E65100"
C_BLUE     = "0D47A1"
C_WHITE    = "FFFFFF"
C_DARK_RED = "B71C1C"
C_DARK_GRN = "1B5E20"

STATUS_LABEL = {"COMPLETED": "DB CONNECTED", "FAILED": "DB NOT CONNECTED", "SKIPPED": "SKIPPED"}
STATUS_COLOR = {"COMPLETED": C_GREEN,        "FAILED": C_RED,              "SKIPPED": C_AMBER}
ERR_LABEL    = {
    "TCP_TIMEOUT": "TCP Connection Timeout",
    "PASSWORD":    "Password Expired / Invalid Username or Password",
    "UNKNOWN":     "Unknown / Other Error",
    "":            "",
}
ERR_COLOR = {"TCP_TIMEOUT": C_RED, "PASSWORD": C_AMBER, "UNKNOWN": C_BLUE, "": ""}


def _border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border():
    s = Side(style="medium", color="90A4AE")
    return Border(left=s, right=s, top=s, bottom=s)


def _ist(dt) -> str:
    return (dt + IST).strftime("%d-%b-%Y %I:%M:%S %p IST") if dt else ""


def _col(n):
    return get_column_letter(n)


# ── Expiry validation helpers ─────────────────────────────────────────────────

def _is_expiry_set(expiry_val) -> bool:
    """Return True if EXPIRY_DATE has a real date value (not null/none/empty)."""
    if expiry_val is None:
        return False
    s = str(expiry_val).strip().upper()
    return s not in ("", "NULL", "NONE", "NONE")


def _compute_expiry_status(fmw_query_rows: list) -> tuple:
    """
    Given list of FMW query result rows, determine expiry status at
    Customer/Environment level.

    Returns (expired_str, action_required_str, reason_str).
    """
    if not fmw_query_rows:
        # No rows returned — treat as no expiry
        return "No", "No", ""

    for row in fmw_query_rows:
        expiry = row.get("EXPIRY_DATE", "")
        if _is_expiry_set(expiry):
            # At least one account has a date — flag the whole env
            return "Yes", "Yes", str(expiry).strip()

    return "No", "No", ""


# ── Key: (customer_name, env_type) → expiry status ───────────────────────────

def _build_expiry_map(results: list) -> dict:
    """
    Build a map of (customer_name, env_type) → (expired, action_required, reason)
    by aggregating fmw_query_rows across all DB entries for that Customer/Env.
    Only processes FMW db_type entries.
    """
    # Collect all FMW query rows per (customer, env)
    env_rows: dict = {}
    for r in results:
        if r.get("db_type", "").upper() != "FMW":
            continue
        key = (r.get("customer_name", ""), r.get("env_type", ""))
        fmw_rows = r.get("fmw_query_rows", [])
        if key not in env_rows:
            env_rows[key] = []
        env_rows[key].extend(fmw_rows)

    expiry_map = {}
    for key, rows in env_rows.items():
        expiry_map[key] = _compute_expiry_status(rows)

    return expiry_map


# ── Sheet 1: DB Connection Check ─────────────────────────────────────────────

def _write_connection_sheet(ws, results, run_id, timestamp, db_label, expiry_map):
    total   = len(results)
    success = sum(1 for r in results if r.get("connection_status") == "COMPLETED")
    failed  = sum(1 for r in results if r.get("connection_status") == "FAILED")
    skipped = sum(1 for r in results if r.get("connection_status") == "SKIPPED")
    tcp_err = sum(1 for r in results if r.get("error_type") == "TCP_TIMEOUT")
    pwd_err = sum(1 for r in results if r.get("error_type") == "PASSWORD")
    unk_err = sum(1 for r in results
                  if r.get("error_type") == "UNKNOWN"
                  and r.get("connection_status") == "FAILED")

    if total == 0:
        summary, summ_color = "No data processed", C_RED
    elif failed == 0 and skipped == 0:
        summary, summ_color = "ALL CONNECTED  ✅", C_GREEN
    elif success == 0:
        summary, summ_color = f"NO SUCCESSFUL CONNECTIONS  ❌   Not Connected: {failed}  |  Skipped: {skipped}", C_RED
    else:
        summary, summ_color = f"PARTIAL  ⚠    Connected: {success}  |  Not Connected: {failed}  |  Skipped: {skipped}", C_AMBER

    last_letter = _col(LAST_COL)

    # Row 1: Title
    ws.merge_cells(f"A1:{last_letter}1")
    c = ws["A1"]
    c.value     = f"{db_label} Oracle DB Connection Check  |  Run #{run_id}  |  {timestamp}"
    c.font      = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=HDR_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: Summary bar
    ws.merge_cells(f"A2:{last_letter}2")
    c = ws["A2"]
    err_part = (f"  |  Errors — TCP: {tcp_err}  |  Password: {pwd_err}  |  Unknown: {unk_err}"
                if failed else "")
    c.value     = (f"Total: {total}   ✅ Connected: {success}   ❌ Not Connected: {failed}"
                   f"   ⚠ Skipped: {skipped}{err_part}   —   {summary}")
    c.font      = Font(name="Calibri", size=10, bold=True, color=summ_color)
    c.fill      = PatternFill("solid", fgColor=SUMM_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Row 3: Headers
    headers = [
        "#", "Customer", "Environment", "Server Name", "Server ID",
        "Server Host", "DB Name", "DB Type", "DSN (Host:Port/Service)",
        "DB Connection Status", "Duration (ms)", "Error", "Error Type",
        "Checked At (IST)",
        "Expired\n(Action Required)", "Reason\n(if yes)",
    ]
    widths = [4, 22, 13, 18, 10, 22, 20, 11, 32, 20, 13, 45, 35, 24, 18, 35]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c           = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
        c.fill      = PatternFill("solid", fgColor=HDR_MED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[_col(col)].width = w
    ws.row_dimensions[3].height = 30

    ws.auto_filter.ref = f"A3:{last_letter}{3 + max(total, 1)}"

    # Data rows
    for idx, r in enumerate(results, 1):
        row      = idx + 3
        stat     = r.get("connection_status", "SKIPPED")
        err_type = r.get("error_type", "") or ""

        bg = (GREEN_BG if idx % 2 else GREEN_BG2) if stat == "COMPLETED" else \
             (RED_BG   if idx % 2 else RED_BG2)   if stat == "FAILED"    else \
             (AMBER_BG if idx % 2 else AMBER_BG2)

        h   = r.get("db_host", "") or ""
        p   = r.get("db_port", 1521) or 1521
        svc = r.get("db_service", "") or ""
        dsn = f"{h}:{p}/{svc}" if h else ""

        # Expiry columns — only for FMW db_type
        db_type_val = r.get("db_type", "").upper()
        if db_type_val == "FMW":
            key = (r.get("customer_name", ""), r.get("env_type", ""))
            expired_str, action_str, reason_str = expiry_map.get(key, ("No", "No", ""))
            expiry_display = f"{expired_str} / {action_str}"
        else:
            expiry_display = ""
            reason_str     = ""

        values = [
            idx,
            r.get("customer_name",  "") or "",
            r.get("env_type",       "") or "",
            r.get("server_name",    "") or "",
            r.get("server_id",      "") or "",
            r.get("server_host",    "") or "",
            r.get("db_name",        "") or "—",
            r.get("db_type",        "") or "",
            dsn,
            STATUS_LABEL.get(stat, stat),
            r.get("duration_ms",    0),
            r.get("error_message",  "") or "",
            ERR_LABEL.get(err_type, err_type),
            _ist(r.get("checked_at")),
            expiry_display,
            reason_str,
        ]

        for col, val in enumerate(values, 1):
            c           = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Calibri", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = _border()
            c.alignment = Alignment(vertical="center", wrap_text=True)

            if col == 10:  # DB Connection Status
                color = STATUS_COLOR.get(stat, C_RED)
                c.font      = Font(name="Calibri", size=10, bold=True, color=color)
                c.alignment = Alignment(horizontal="center", vertical="center")

            if col == 11:  # Duration — right-align
                c.alignment = Alignment(horizontal="right", vertical="center")

            if col == 12 and val:  # Error — italic
                c.font = Font(name="Calibri", size=9, italic=True,
                              color=STATUS_COLOR.get(stat, C_RED))

            if col == 13 and val:  # Error Type — bold, typed colour
                c.font = Font(name="Calibri", size=9, bold=True,
                              color=ERR_COLOR.get(err_type, C_BLUE))

            if col == 15 and db_type_val == "FMW":  # Expired / Action Required
                if expiry_display.startswith("Yes"):
                    c.font = Font(name="Calibri", size=10, bold=True, color=C_DARK_RED)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.font = Font(name="Calibri", size=10, bold=True, color=C_DARK_GRN)
                    c.alignment = Alignment(horizontal="center", vertical="center")

            if col == 16 and reason_str:  # Reason
                c.font = Font(name="Calibri", size=9, italic=True, color=C_DARK_RED)

        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A4"


# ── Sheet 2: Customer - Environment Report ────────────────────────────────────

def _write_fmw_report_sheet(ws, results, run_id, timestamp, db_label):
    """
    Write FMW query results to the Customer - Environment Report sheet.
    One row per OAM account per Customer/Environment (FMW db_type only).
    Non-FMW entries are completely excluded.
    """
    SHEET_COLS  = 6
    last_letter = _col(SHEET_COLS)

    # Row 1: Title
    ws.merge_cells(f"A1:{last_letter}1")
    c = ws["A1"]
    c.value     = f"Customer - Environment Report  |  FMW OAM Account Expiry  |  Run #{run_id}  |  {timestamp}"
    c.font      = Font(name="Calibri", size=13, bold=True, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=HDR_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: Sub-title
    ws.merge_cells(f"A2:{last_letter}2")
    c = ws["A2"]
    c.value     = ("FMW database type only  |  "
                   "Query: SELECT USERNAME, ACCOUNT_STATUS, EXPIRY_DATE FROM DBA_USERS WHERE USERNAME LIKE '%OAM%'")
    c.font      = Font(name="Calibri", size=9, italic=True, color="37474F")
    c.fill      = PatternFill("solid", fgColor="E8F5E9")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3: Headers
    headers = [
        "Customer", "Environment", "Server Host / DB Name / DB Type",
        "USERNAME", "ACCOUNT_STATUS", "EXPIRY_DATE",
    ]
    widths = [22, 14, 38, 26, 20, 22]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c           = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
        c.fill      = PatternFill("solid", fgColor=HDR_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[_col(col)].width = w
    ws.row_dimensions[3].height = 22

    # Gather FMW rows
    data_rows = []
    for r in results:
        if r.get("db_type", "").upper() != "FMW":
            continue
        fmw_rows = r.get("fmw_query_rows", [])
        server_host = r.get("server_host", "") or ""
        db_name     = r.get("db_name",     "") or ""
        db_type     = r.get("db_type",     "") or ""
        conn_info   = f"{server_host} / {db_name} / {db_type}"
        cname       = r.get("customer_name", "") or ""
        env_type    = r.get("env_type",      "") or ""

        if fmw_rows:
            for frow in fmw_rows:
                data_rows.append({
                    "customer":    cname,
                    "env_type":    env_type,
                    "conn_info":   conn_info,
                    "USERNAME":    frow.get("USERNAME",       ""),
                    "ACCOUNT_STATUS": frow.get("ACCOUNT_STATUS", ""),
                    "EXPIRY_DATE": frow.get("EXPIRY_DATE",    ""),
                })
        else:
            # FMW connection existed but query returned no rows (or connection failed)
            data_rows.append({
                "customer":       cname,
                "env_type":       env_type,
                "conn_info":      conn_info,
                "USERNAME":       "",
                "ACCOUNT_STATUS": "",
                "EXPIRY_DATE":    "",
            })

    ws.auto_filter.ref = f"A3:{last_letter}{3 + max(len(data_rows), 1)}"

    for idx, dr in enumerate(data_rows, 1):
        row      = idx + 3
        bg       = "E8F5E9" if idx % 2 else "F1F8E9"
        expiry   = dr.get("EXPIRY_DATE", "")
        has_date = _is_expiry_set(expiry)

        values = [
            dr["customer"],
            dr["env_type"],
            dr["conn_info"],
            dr["USERNAME"],
            dr["ACCOUNT_STATUS"],
            expiry,
        ]

        for col, val in enumerate(values, 1):
            c           = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Calibri", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = _border()
            c.alignment = Alignment(vertical="center", wrap_text=True)

            # Highlight EXPIRY_DATE if it has an actual date
            if col == 6 and has_date:
                c.font = Font(name="Calibri", size=10, bold=True, color=C_DARK_RED)
                c.alignment = Alignment(horizontal="center", vertical="center")

            # Highlight ACCOUNT_STATUS
            if col == 5 and val:
                if "EXPIRED" in str(val).upper() or "LOCKED" in str(val).upper():
                    c.font = Font(name="Calibri", size=10, bold=True, color=C_RED)
                else:
                    c.font = Font(name="Calibri", size=10, color=C_DARK_GRN)

        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A4"

    if not data_rows:
        ws.merge_cells(f"A4:{last_letter}4")
        c = ws["A4"]
        c.value     = "No FMW database connections found for this run."
        c.font      = Font(name="Calibri", size=10, italic=True, color=C_AMBER)
        c.alignment = Alignment(horizontal="center", vertical="center")


# ── Sheet name helpers ────────────────────────────────────────────────────────

def _safe_sheet_name(name: str) -> str:
    """Sanitise to a valid Excel sheet name (max 31 chars, no illegal chars)."""
    for ch in r"\/:*?[]":
        name = name.replace(ch, " ")
    return name[:31].strip()


def _write_fmw_env_sheet(ws, cname: str, env_type: str,
                         env_results: list, run_id, timestamp):
    """
    Write one sheet for a single Customer + Environment.
    env_results = list of result dicts (already filtered to this cname/env_type, FMW only).
    Columns: Server Host/DB/Type | USERNAME | ACCOUNT_STATUS | EXPIRY_DATE | Action Required
    """
    SHEET_COLS  = 5
    last_letter = _col(SHEET_COLS)

    # Row 1: Title
    ws.merge_cells(f"A1:{last_letter}1")
    c = ws["A1"]
    c.value     = f"{cname}  |  {env_type}  |  FMW OAM Account Expiry  |  Run #{run_id}  |  {timestamp}"
    c.font      = Font(name="Calibri", size=12, bold=True, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=HDR_GREEN)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Row 2: Sub-title / query reference
    ws.merge_cells(f"A2:{last_letter}2")
    c = ws["A2"]
    c.value = ("FMW database type only  |  "
               "Query: SELECT USERNAME, ACCOUNT_STATUS, EXPIRY_DATE FROM DBA_USERS WHERE USERNAME LIKE '%OAM%'")
    c.font      = Font(name="Calibri", size=9, italic=True, color="37474F")
    c.fill      = PatternFill("solid", fgColor="E8F5E9")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Row 3: Headers
    headers = ["Server Host / DB Name / DB Type", "USERNAME", "ACCOUNT_STATUS", "EXPIRY_DATE", "Action Required"]
    widths  = [40, 26, 20, 22, 18]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c           = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Calibri", size=10, bold=True, color=C_WHITE)
        c.fill      = PatternFill("solid", fgColor=HDR_GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
        ws.column_dimensions[_col(col)].width = w
    ws.row_dimensions[3].height = 22

    # Flatten all OAM account rows from every DB entry for this Customer/Env
    data_rows = []
    for r in env_results:
        conn_info = (f"{r.get('server_host','') or ''} / "
                     f"{r.get('db_name','') or ''} / "
                     f"{r.get('db_type','') or ''}")
        fmw_rows = r.get("fmw_query_rows", [])

        if fmw_rows:
            for frow in fmw_rows:
                expiry     = frow.get("EXPIRY_DATE", "") or ""
                expiry_set = _is_expiry_set(expiry)
                data_rows.append({
                    "conn_info":      conn_info,
                    "USERNAME":       frow.get("USERNAME",       ""),
                    "ACCOUNT_STATUS": frow.get("ACCOUNT_STATUS", ""),
                    "EXPIRY_DATE":    expiry,
                    "action":         "Yes" if expiry_set else "No",
                })
        else:
            data_rows.append({
                "conn_info": conn_info, "USERNAME": "",
                "ACCOUNT_STATUS": "", "EXPIRY_DATE": "", "action": "No",
            })

    ws.auto_filter.ref = f"A3:{last_letter}{3 + max(len(data_rows), 1)}"

    for idx, dr in enumerate(data_rows, 1):
        row      = idx + 3
        bg       = "E8F5E9" if idx % 2 else "F1F8E9"
        expiry   = dr["EXPIRY_DATE"]
        has_date = _is_expiry_set(expiry)
        action   = dr["action"]

        for col, val in enumerate(
            [dr["conn_info"], dr["USERNAME"], dr["ACCOUNT_STATUS"], expiry, action], 1
        ):
            c           = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name="Calibri", size=10)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = _border()
            c.alignment = Alignment(vertical="center", wrap_text=True)

            if col == 4 and has_date:   # EXPIRY_DATE with value → red bold
                c.font      = Font(name="Calibri", size=10, bold=True, color=C_DARK_RED)
                c.alignment = Alignment(horizontal="center", vertical="center")

            if col == 3 and val:        # ACCOUNT_STATUS
                if "EXPIRED" in str(val).upper() or "LOCKED" in str(val).upper():
                    c.font = Font(name="Calibri", size=10, bold=True, color=C_RED)
                else:
                    c.font = Font(name="Calibri", size=10, color=C_DARK_GRN)

            if col == 5:                # Action Required
                c.font      = Font(name="Calibri", size=10, bold=True,
                                   color=C_DARK_RED if action == "Yes" else C_DARK_GRN)
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row].height = 20

    ws.freeze_panes = "A4"

    if not data_rows:
        ws.merge_cells(f"A4:{last_letter}4")
        c = ws["A4"]
        c.value     = f"No FMW OAM accounts found for {cname} / {env_type}."
        c.font      = Font(name="Calibri", size=10, italic=True, color=C_AMBER)
        c.alignment = Alignment(horizontal="center", vertical="center")


# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel(results, run_id, timestamp, excel_path, db_label="FMW") -> str:
    os.makedirs(excel_path, exist_ok=True)

    label    = (db_label or "FMW").upper().replace("/", "_").replace(" ", "_")
    date_str = datetime.now().strftime("%d-%m-%y")
    filename = f"{label}_CONNECTION_CHECKS_{date_str}-{run_id}.xlsx"
    filepath = os.path.join(excel_path, filename)

    expiry_map = _build_expiry_map(results)

    wb = Workbook()

    # ── Sheet 1: DB Connection Check (unchanged) ──────────────────────────────
    ws1 = wb.active
    ws1.title = "DB Connection Check"
    _write_connection_sheet(ws1, results, run_id, timestamp, db_label, expiry_map)

    # ── Dynamic sheets: one per unique Customer + Environment (FMW only) ──────
    # Preserves the order results appear in; groups all DB entries per pair.
    env_groups: dict = {}   # (customer_name, env_type) → [result, ...]
    for r in results:
        if r.get("db_type", "").upper() != "FMW":
            continue
        key = (r.get("customer_name", "") or "", r.get("env_type", "") or "")
        env_groups.setdefault(key, []).append(r)

    used_names: dict = {}   # track used sheet names to avoid duplicates

    for (cname, env_type), env_results in env_groups.items():
        raw_name  = f"{cname} - {env_type}"
        safe_name = _safe_sheet_name(raw_name)

        # If two customers produce the same safe name, append a counter
        if safe_name in used_names:
            used_names[safe_name] += 1
            safe_name = _safe_sheet_name(f"{safe_name} ({used_names[safe_name]})")
        else:
            used_names[safe_name] = 1

        ws = wb.create_sheet(title=safe_name)
        _write_fmw_env_sheet(ws, cname, env_type, env_results, run_id, timestamp)
        log.info(f"  FMW sheet   : '{safe_name}'  ({len(env_results)} DB entry(s))")

    wb.save(filepath)
    log.info(f"  Excel saved : {filepath}")
    return filepath
